from datetime import date
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Colors
VERD_FOSC = "2E7D32"
TARONJA_FOSC = "E65100"
GRIS_CLAR = "F5F5F5"
GRIS_TOTAL = "BDBDBD"
BLANC = "FFFFFF"


def _aplicar_estil_cap(cel, bg_color: str, text_color: str = BLANC):
    cel.font = Font(bold=True, color=text_color, size=11)
    cel.fill = PatternFill("solid", fgColor=bg_color)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _vora_fina(cel)


def _vora_fina(cel):
    fina = Side(style="thin", color="CCCCCC")
    cel.border = Border(left=fina, right=fina, top=fina, bottom=fina)


def _ajustar_amplades(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cel in col:
            try:
                val = str(cel.value) if cel.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def _full_comandes_per_client(wb, comandes: list):
    ws = wb.active
    ws.title = "Comandes per Client"
    ws.freeze_panes = "A2"

    headers = ["Client", "Telèfon", "Article", "Quantitat", "Unitat", "Hora"]
    for col_idx, h in enumerate(headers, 1):
        cel = ws.cell(row=1, column=col_idx, value=h)
        _aplicar_estil_cap(cel, VERD_FOSC)

    fila = 2
    for comanda in comandes:
        client = comanda.get("client", "Desconegut")
        telefon = comanda.get("telefon", "")
        hora = comanda.get("hora", "")
        articles = comanda.get("articles", [])

        for i, article in enumerate(articles):
            bg = BLANC if fila % 2 == 0 else GRIS_CLAR
            fill = PatternFill("solid", fgColor=bg)

            valors = [
                client if i == 0 else "",
                telefon if i == 0 else "",
                article.get("nom", ""),
                article.get("quantitat", 0),
                article.get("unitat", ""),
                hora if i == 0 else "",
            ]

            for col_idx, val in enumerate(valors, 1):
                cel = ws.cell(row=fila, column=col_idx, value=val)
                cel.fill = fill
                cel.alignment = Alignment(vertical="center")
                _vora_fina(cel)
                if col_idx == 4:  # Quantitat
                    cel.number_format = "0.##"

            fila += 1

        if not articles:
            bg = BLANC if fila % 2 == 0 else GRIS_CLAR
            fill = PatternFill("solid", fgColor=bg)
            for col_idx, val in enumerate([client, telefon, "", "", "", hora], 1):
                cel = ws.cell(row=fila, column=col_idx, value=val)
                cel.fill = fill
                cel.alignment = Alignment(vertical="center")
                _vora_fina(cel)
            fila += 1

    ws.row_dimensions[1].height = 20
    _ajustar_amplades(ws)


def _full_resum_total(wb, comandes: list):
    ws = wb.create_sheet(title="Resum Total del Dia")
    ws.freeze_panes = "A2"

    headers = ["Article", "Quantitat Total", "Unitat", "Clients"]
    for col_idx, h in enumerate(headers, 1):
        cel = ws.cell(row=1, column=col_idx, value=h)
        _aplicar_estil_cap(cel, TARONJA_FOSC)

    # Agrupar per (nom_article, unitat)
    agrupat = defaultdict(lambda: {"quantitat": 0.0, "clients": []})

    for comanda in comandes:
        client = comanda.get("client", "Desconegut")
        for article in comanda.get("articles", []):
            clau = (article.get("nom", "").lower(), article.get("unitat", ""))
            agrupat[clau]["quantitat"] += float(article.get("quantitat", 0))
            if client not in agrupat[clau]["clients"]:
                agrupat[clau]["clients"].append(client)

    articles_ordenats = sorted(agrupat.items(), key=lambda x: x[0][0])

    fila = 2
    for (nom, unitat), dades in articles_ordenats:
        bg = BLANC if fila % 2 == 0 else GRIS_CLAR
        fill = PatternFill("solid", fgColor=bg)

        valors = [
            nom,
            dades["quantitat"],
            unitat,
            ", ".join(dades["clients"]),
        ]

        for col_idx, val in enumerate(valors, 1):
            cel = ws.cell(row=fila, column=col_idx, value=val)
            cel.fill = fill
            cel.alignment = Alignment(vertical="center")
            _vora_fina(cel)
            if col_idx == 2:
                cel.number_format = "0.##"

        fila += 1

    # Fila de total
    if fila > 2:
        fill_total = PatternFill("solid", fgColor=GRIS_TOTAL)
        font_total = Font(bold=True, size=11)

        ws.cell(row=fila, column=1, value="TOTAL ARTICLES").font = font_total
        ws.cell(row=fila, column=1).fill = fill_total
        ws.cell(row=fila, column=1).alignment = Alignment(horizontal="right", vertical="center")
        _vora_fina(ws.cell(row=fila, column=1))

        total_cel = ws.cell(row=fila, column=2)
        total_cel.value = f"=SUM(B2:B{fila - 1})"
        total_cel.font = font_total
        total_cel.fill = fill_total
        total_cel.number_format = "0.##"
        total_cel.alignment = Alignment(vertical="center")
        _vora_fina(total_cel)

        for col_idx in [3, 4]:
            cel = ws.cell(row=fila, column=col_idx)
            cel.fill = fill_total
            _vora_fina(cel)

    ws.row_dimensions[1].height = 20
    _ajustar_amplades(ws)


def generar_excel(comandes: list, nom_fitxer: str = None) -> str:
    """
    Rep la llista de comandes d'avui.
    Retorna el path del fitxer Excel generat.
    """
    if nom_fitxer is None:
        nom_fitxer = f"comandes_{date.today().isoformat()}.xlsx"

    output_dir = Path(__file__).parent / "data"
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / nom_fitxer

    wb = Workbook()
    _full_comandes_per_client(wb, comandes)
    _full_resum_total(wb, comandes)
    wb.save(str(path))

    return str(path)
